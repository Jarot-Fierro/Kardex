from datetime import date, datetime

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def export_queryset_to_excel(queryset, filename='reporte', excluded_fields=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    model = queryset.model
    opts = model._meta
    fields = [
        f for f in opts.concrete_fields
        if f.name not in (excluded_fields or [])
    ]

    # === TÍTULO ===
    title = opts.verbose_name_plural.title()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # === ENCABEZADOS ===
    headers = [f.verbose_name.title() for f in fields]
    ws.append([])  # fila vacía después del título
    ws.append(headers)

    # === FILAS ===
    for obj in queryset:
        row = []
        for field in fields:
            # Si el campo tiene choices → usar display
            if field.choices:
                value = getattr(obj, f"get_{field.name}_display")()
            else:
                value = getattr(obj, field.name)

            # Manejo de tipos
            if value is None:
                row.append('')
            elif isinstance(value, str):
                row.append(value)
            elif isinstance(value, (int, float, bool)):
                row.append(value)
            elif isinstance(value, (date, datetime)):
                fmt = "%d-%m-%Y %H:%M" if isinstance(value, datetime) else "%d-%m-%Y"
                row.append(value.strftime(fmt))
            else:
                row.append(str(value))

        ws.append(row)

    # === AUTOAJUSTE DE ANCHO DE COLUMNAS ===
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(i)  # más seguro que usar col[0]
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # === RESPUESTA ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    wb.save(response)
    return response


def export_generic_history_to_excel(queryset, filename='historial', sheet_title='Historial'):
    if not queryset.exists():
        raise ValueError("El queryset está vacío.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Obtener campos base del modelo
    model = queryset.model
    base_fields = [f.name for f in model._meta.concrete_fields]

    # Tomar un ejemplo de data para inferir campos JSON (si existe)
    example_entry = queryset.first()
    json_data_keys = list(example_entry.data.keys()) if hasattr(example_entry, 'data') else []

    # Combinar encabezados
    headers = base_fields + json_data_keys
    ws.append(headers)

    # Escribir filas
    for entry in queryset:
        base_values = []
        for field in base_fields:
            value = getattr(entry, field)
            base_values.append(str(value) if value is not None else '')

        data_values = []
        for key in json_data_keys:
            value = entry.data.get(key, '')
            data_values.append(str(value))

        ws.append(base_values + data_values)

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    wb.save(response)
    return response


def export_extended_queryset_to_excel(queryset, fields, filename='reporte', sheet_title='Datos'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Escribir encabezados
    ws.append([field[0] for field in fields])  # Ejemplo: ("Código", "code")

    # Escribir filas
    for obj in queryset:
        row = []
        for label, key in fields:
            value = None

            if '__' in key:
                # Soporta relaciones como support_info__problem
                parts = key.split('__')
                current = obj
                for part in parts:
                    current = getattr(current, part, None)
                    if current is None:
                        break
                value = str(current) if current is not None else ''
            else:
                # Campos directos de Transaction u otros objetos
                if hasattr(obj, key):
                    value = getattr(obj, key)
                elif hasattr(obj, 'support_info') and key.startswith('support_info_'):
                    sub_key = key.replace('support_info_', '')
                    value = getattr(obj.support_info, sub_key, '')
                elif hasattr(obj, 'output_info') and key.startswith('output_info_'):
                    sub_key = key.replace('output_info_', '')
                    value = getattr(obj.output_info, sub_key, '')
                else:
                    value = ''

            row.append(str(value) if value is not None else '')

        ws.append(row)

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    wb.save(response)
    return response
